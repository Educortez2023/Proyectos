import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.models.ubicacion_model import UbicacionModel


class VentanaUbicaciones(ctk.CTkToplevel):

    def __init__(self, parent):

        super().__init__(parent)

        # =====================================================
        # CONFIGURACIÓN GENERAL
        # =====================================================

        self.parent = parent

        self.title("Gestión de Ubicaciones")

        self.geometry("1100x650")
        self.minsize(900, 550)

        self.transient(parent)

        # =====================================================
        # MODELO
        # =====================================================

        self.model = UbicacionModel()

        # =====================================================
        # REFERENCIA AL FORMULARIO
        # =====================================================

        self.ventana_formulario = None

        # =====================================================
        # CREAR INTERFAZ
        # =====================================================

        self.crear_interfaz()

        # =====================================================
        # CARGAR DATOS
        # =====================================================

        self.cargar_ubicaciones()

        # =====================================================
        # EVENTOS
        # =====================================================

        self.bind("<Configure>", self.on_resize)

        # =====================================================
        # MOSTRAR VENTANA
        # =====================================================

        self.after(150, self.mostrar_delante)

    # =========================================================
    # MOSTRAR VENTANA DELANTE
    # =========================================================

    def mostrar_delante(self):

        try:
            self.lift()
            self.focus_force()

        except Exception:
            pass

    # =========================================================
    # DETECTAR CAMBIO DE TAMAÑO
    # =========================================================

    def on_resize(self, event=None):

        try:

            if not hasattr(self, "tabla"):
                return

            if event is not None and event.widget != self:
                return

            ancho = self.tabla.winfo_width()

            if ancho <= 1:
                return

            self.ajustar_columnas(ancho)

        except Exception:
            pass

    # =========================================================
    # AJUSTAR COLUMNAS
    # =========================================================

    def ajustar_columnas(self, ancho=None):

        try:

            if not hasattr(self, "tabla"):
                return

            if ancho is None:
                ancho = self.tabla.winfo_width()

            if ancho <= 1:
                return

            # -------------------------------------------------
            # Ancho mínimo total de la tabla
            # -------------------------------------------------

            ancho_minimo_total = (
                60 +
                180 +
                300 +
                100
            )

            # -------------------------------------------------
            # Si la ventana es demasiado pequeña,
            # mantenemos los mínimos para permitir
            # desplazamiento horizontal.
            # -------------------------------------------------

            if ancho < ancho_minimo_total:

                self.tabla.column(
                    "id",
                    width=60
                )

                self.tabla.column(
                    "nombre",
                    width=180
                )

                self.tabla.column(
                    "descripcion",
                    width=300
                )

                self.tabla.column(
                    "activo",
                    width=100
                )

                return

            # -------------------------------------------------
            # Espacio disponible
            # -------------------------------------------------

            espacio = ancho - 10

            # -------------------------------------------------
            # Distribución proporcional
            # -------------------------------------------------

            proporciones = {
                "id": 0.08,
                "nombre": 0.25,
                "descripcion": 0.52,
                "activo": 0.15
            }

            minimos = {
                "id": 60,
                "nombre": 150,
                "descripcion": 220,
                "activo": 90
            }

            for columna, proporcion in proporciones.items():

                nuevo_ancho = int(
                    espacio * proporcion
                )

                if nuevo_ancho < minimos[columna]:
                    nuevo_ancho = minimos[columna]

                self.tabla.column(
                    columna,
                    width=nuevo_ancho
                )

        except Exception:
            pass

    # =========================================================
    # CREAR INTERFAZ
    # =========================================================

    def crear_interfaz(self):

        # =====================================================
        # CONFIGURACIÓN PRINCIPAL
        # =====================================================

        self.grid_rowconfigure(
            2,
            weight=1
        )

        self.grid_columnconfigure(
            0,
            weight=1
        )

        # =====================================================
        # TÍTULO
        # =====================================================

        titulo = ctk.CTkLabel(
            self,
            text="Gestión de Ubicaciones",
            font=("Arial", 24, "bold")
        )

        titulo.grid(
            row=0,
            column=0,
            padx=20,
            pady=(20, 10),
            sticky="w"
        )

        # =====================================================
        # FRAME DE CONTROLES
        # =====================================================

        frame_controles = ctk.CTkFrame(
            self
        )

        frame_controles.grid(
            row=1,
            column=0,
            padx=20,
            pady=(5, 10),
            sticky="ew"
        )

        frame_controles.grid_columnconfigure(
            5,
            weight=1
        )

        # =====================================================
        # BOTÓN NUEVO
        # =====================================================

        btn_nuevo = ctk.CTkButton(
            frame_controles,
            text="➕ Nueva Ubicación",
            width=160,
            height=36,
            command=self.abrir_nuevo
        )

        btn_nuevo.grid(
            row=0,
            column=0,
            padx=5,
            pady=10
        )

        # =====================================================
        # BOTÓN EDITAR
        # =====================================================

        btn_editar = ctk.CTkButton(
            frame_controles,
            text="✏ Editar",
            width=110,
            height=36,
            command=self.editar_ubicacion
        )

        btn_editar.grid(
            row=0,
            column=1,
            padx=5,
            pady=10
        )

        # =====================================================
        # BOTÓN ELIMINAR
        # =====================================================

        btn_eliminar = ctk.CTkButton(
            frame_controles,
            text="🗑 Eliminar",
            width=120,
            height=36,
            command=self.eliminar_ubicacion
        )

        btn_eliminar.grid(
            row=0,
            column=2,
            padx=5,
            pady=10
        )

        # =====================================================
        # BOTÓN ACTUALIZAR
        # =====================================================

        btn_actualizar = ctk.CTkButton(
            frame_controles,
            text="🔄 Actualizar",
            width=120,
            height=36,
            command=self.cargar_ubicaciones
        )

        btn_actualizar.grid(
            row=0,
            column=3,
            padx=5,
            pady=10
        )

        # =====================================================
        # BOTÓN EXPORTAR
        # =====================================================

        btn_exportar = ctk.CTkButton(
            frame_controles,
            text="📊 Exportar a Excel",
            width=150,
            height=36,
            command=self.exportar_excel
        )

        btn_exportar.grid(
            row=0,
            column=4,
            padx=5,
            pady=10
        )

        # =====================================================
        # CAMPO DE BÚSQUEDA
        # =====================================================

        self.entrada_busqueda = ctk.CTkEntry(
            frame_controles,
            placeholder_text="Buscar ubicación...",
            height=36
        )

        self.entrada_busqueda.grid(
            row=0,
            column=5,
            padx=(10, 10),
            pady=10,
            sticky="ew"
        )

        # =====================================================
        # BUSCAR MIENTRAS ESCRIBE
        # =====================================================

        self.entrada_busqueda.bind(
            "<KeyRelease>",
            self.buscar_ubicaciones
        )

        # =====================================================
        # FRAME TABLA
        # =====================================================

        frame_tabla = ctk.CTkFrame(
            self
        )

        frame_tabla.grid(
            row=2,
            column=0,
            padx=20,
            pady=(0, 20),
            sticky="nsew"
        )

        # =====================================================
        # CONFIGURACIÓN GRID
        # =====================================================

        frame_tabla.grid_rowconfigure(
            0,
            weight=1
        )

        frame_tabla.grid_columnconfigure(
            0,
            weight=1
        )

        # =====================================================
        # SCROLLBAR VERTICAL
        # =====================================================

        scrollbar_vertical = ttk.Scrollbar(
            frame_tabla,
            orient="vertical"
        )

        scrollbar_vertical.grid(
            row=0,
            column=1,
            sticky="ns"
        )

        # =====================================================
        # SCROLLBAR HORIZONTAL
        # =====================================================

        scrollbar_horizontal = ttk.Scrollbar(
            frame_tabla,
            orient="horizontal"
        )

        scrollbar_horizontal.grid(
            row=1,
            column=0,
            sticky="ew"
        )

        # =====================================================
        # COLUMNAS
        # =====================================================

        columnas = (
            "id",
            "nombre",
            "descripcion",
            "activo"
        )

        # =====================================================
        # TREEVIEW
        # =====================================================

        self.tabla = ttk.Treeview(
            frame_tabla,
            columns=columnas,
            show="headings",
            yscrollcommand=scrollbar_vertical.set,
            xscrollcommand=scrollbar_horizontal.set
        )

        self.tabla.grid(
            row=0,
            column=0,
            sticky="nsew"
        )

        # =====================================================
        # ENCABEZADOS
        # =====================================================

        self.tabla.heading(
            "id",
            text="ID"
        )

        self.tabla.heading(
            "nombre",
            text="Nombre"
        )

        self.tabla.heading(
            "descripcion",
            text="Descripción"
        )

        self.tabla.heading(
            "activo",
            text="Estado"
        )

        # =====================================================
        # CONFIGURACIÓN DE COLUMNAS
        # =====================================================

        self.tabla.column(
            "id",
            width=60,
            minwidth=60,
            anchor="center",
            stretch=False
        )

        self.tabla.column(
            "nombre",
            width=220,
            minwidth=150,
            anchor="w",
            stretch=True
        )

        self.tabla.column(
            "descripcion",
            width=400,
            minwidth=220,
            anchor="w",
            stretch=True
        )

        self.tabla.column(
            "activo",
            width=120,
            minwidth=90,
            anchor="center",
            stretch=False
        )

        # =====================================================
        # CONECTAR SCROLLBAR VERTICAL
        # =====================================================

        scrollbar_vertical.config(
            command=self.tabla.yview
        )

        # =====================================================
        # CONECTAR SCROLLBAR HORIZONTAL
        # =====================================================

        scrollbar_horizontal.config(
            command=self.tabla.xview
        )

        # =====================================================
        # DOBLE CLIC PARA EDITAR
        # =====================================================

        self.tabla.bind(
            "<Double-1>",
            lambda event: self.editar_ubicacion()
        )

    # =========================================================
    # CARGAR UBICACIONES
    # =========================================================

    def cargar_ubicaciones(self):

        try:

            # =================================================
            # LIMPIAR TABLA
            # =================================================

            for item in self.tabla.get_children():

                self.tabla.delete(
                    item
                )

            # =================================================
            # CONSULTAR BASE DE DATOS
            # =================================================

            ubicaciones = self.model.listar()

            # =================================================
            # INSERTAR DATOS
            # =================================================

            for ubicacion in ubicaciones:

                estado = (
                    "Activo"
                    if ubicacion[3] == 1
                    else "Inactivo"
                )

                self.tabla.insert(
                    "",
                    "end",
                    values=(
                        ubicacion[0],
                        ubicacion[1],
                        ubicacion[2] or "",
                        estado
                    )
                )

            print(
                "Ubicaciones cargadas:",
                len(ubicaciones)
            )

            # =================================================
            # AJUSTAR COLUMNAS DESPUÉS DE CARGAR
            # =================================================

            self.after(
                100,
                self.on_resize
            )

        except Exception as e:

            print(
                "Error al cargar ubicaciones:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudieron cargar "
                    "las ubicaciones:\n\n"
                    f"{e}"
                ),
                parent=self
            )

    # =========================================================
    # BUSCAR UBICACIONES
    # =========================================================

    def buscar_ubicaciones(self, event=None):

        texto = (
            self.entrada_busqueda
            .get()
            .strip()
        )

        try:

            # =================================================
            # LIMPIAR TABLA
            # =================================================

            for item in self.tabla.get_children():

                self.tabla.delete(
                    item
                )

            # =================================================
            # CONSULTAR
            # =================================================

            if texto:

                ubicaciones = self.model.buscar(
                    texto
                )

            else:

                ubicaciones = self.model.listar()

            # =================================================
            # INSERTAR RESULTADOS
            # =================================================

            for ubicacion in ubicaciones:

                estado = (
                    "Activo"
                    if ubicacion[3] == 1
                    else "Inactivo"
                )

                self.tabla.insert(
                    "",
                    "end",
                    values=(
                        ubicacion[0],
                        ubicacion[1],
                        ubicacion[2] or "",
                        estado
                    )
                )

            print(
                "Ubicaciones encontradas:",
                len(ubicaciones)
            )

            self.after(
                50,
                self.on_resize
            )

        except Exception as e:

            print(
                "Error al buscar ubicaciones:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudieron buscar "
                    "las ubicaciones:\n\n"
                    f"{e}"
                ),
                parent=self
            )

    # =========================================================
    # EXPORTAR A EXCEL
    # =========================================================

    def exportar_excel(self):

        try:

            # =================================================
            # OBTENER DATOS DESDE LA BD
            # =================================================

            ubicaciones = self.model.listar()

            if not ubicaciones:

                messagebox.showwarning(
                    "Sin datos",
                    "No existen ubicaciones para exportar.",
                    parent=self
                )

                return

            # =================================================
            # SELECCIONAR ARCHIVO
            # =================================================

            ruta_archivo = filedialog.asksaveasfilename(
                parent=self,
                title="Guardar ubicaciones en Excel",
                defaultextension=".xlsx",
                filetypes=[
                    (
                        "Archivo de Excel",
                        "*.xlsx"
                    ),
                    (
                        "Todos los archivos",
                        "*.*"
                    )
                ],
                initialfile="Reporte_Ubicaciones.xlsx"
            )

            if not ruta_archivo:
                return

            # =================================================
            # CREAR LIBRO
            # =================================================

            libro = Workbook()

            hoja = libro.active

            hoja.title = "Ubicaciones"

            # =================================================
            # ENCABEZADOS
            # =================================================

            encabezados = [
                "ID",
                "Nombre",
                "Descripción",
                "Estado"
            ]

            hoja.append(
                encabezados
            )

            # =================================================
            # FORMATO ENCABEZADOS
            # =================================================

            for celda in hoja[1]:

                celda.font = Font(
                    bold=True
                )

            # =================================================
            # DATOS
            # =================================================

            for ubicacion in ubicaciones:

                estado = (
                    "Activo"
                    if ubicacion[3] == 1
                    else "Inactivo"
                )

                hoja.append(
                    [
                        ubicacion[0],
                        ubicacion[1],
                        ubicacion[2] or "",
                        estado
                    ]
                )

            # =================================================
            # AJUSTAR COLUMNAS
            # =================================================

            for columna in hoja.columns:

                longitud_maxima = 0

                numero_columna = (
                    columna[0].column
                )

                letra_columna = get_column_letter(
                    numero_columna
                )

                for celda in columna:

                    if celda.value is not None:

                        longitud = len(
                            str(celda.value)
                        )

                        if longitud > longitud_maxima:
                            longitud_maxima = longitud

                hoja.column_dimensions[
                    letra_columna
                ].width = min(
                    longitud_maxima + 2,
                    60
                )

            # =================================================
            # CONGELAR ENCABEZADOS
            # =================================================

            hoja.freeze_panes = "A2"

            # =================================================
            # FILTRO
            # =================================================

            hoja.auto_filter.ref = hoja.dimensions

            # =================================================
            # GUARDAR
            # =================================================

            libro.save(
                ruta_archivo
            )

            # =================================================
            # CONFIRMACIÓN
            # =================================================

            messagebox.showinfo(
                "Exportación completada",
                (
                    "Las ubicaciones se exportaron "
                    "correctamente a Excel.\n\n"
                    f"Registros exportados: "
                    f"{len(ubicaciones)}"
                ),
                parent=self
            )

            print(
                "Excel de ubicaciones exportado:",
                ruta_archivo
            )

        except Exception as e:

            print(
                "Error al exportar ubicaciones:",
                e
            )

            messagebox.showerror(
                "Error al exportar",
                (
                    "No se pudieron exportar "
                    "las ubicaciones a Excel:\n\n"
                    f"{e}"
                ),
                parent=self
            )

    # =========================================================
    # NUEVA UBICACIÓN
    # =========================================================

    def abrir_nuevo(self):

        # =================================================
        # EVITAR MÚLTIPLES FORMULARIOS
        # =================================================

        if (
            self.ventana_formulario is not None
            and self.ventana_formulario.winfo_exists()
        ):

            self.ventana_formulario.lift()
            self.ventana_formulario.focus_force()

            return

        try:

            from app.views.formulario_ubicacion import (
                FormularioUbicacion
            )

            # =================================================
            # CREAR FORMULARIO
            # =================================================

            self.ventana_formulario = FormularioUbicacion(
                self,
                callback=self.cargar_ubicaciones
            )

            self.ventana_formulario.transient(
                self
            )

            self.ventana_formulario.lift()
            self.ventana_formulario.focus_force()

            self.ventana_formulario.grab_set()

            # =================================================
            # CERRAR FORMULARIO
            # =================================================

            self.ventana_formulario.protocol(
                "WM_DELETE_WINDOW",
                self.cerrar_formulario
            )

        except Exception as e:

            self.ventana_formulario = None

            print(
                "Error al abrir formulario:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudo abrir el formulario "
                    "de ubicación:\n\n"
                    f"{e}"
                ),
                parent=self
            )

    # =========================================================
    # CERRAR FORMULARIO
    # =========================================================

    def cerrar_formulario(self):

        try:

            if (
                self.ventana_formulario is not None
                and self.ventana_formulario.winfo_exists()
            ):

                try:
                    self.ventana_formulario.grab_release()
                except Exception:
                    pass

                self.ventana_formulario.destroy()

        except Exception as e:

            print(
                "Error al cerrar formulario:",
                e
            )

        finally:

            self.ventana_formulario = None

            try:

                self.lift()
                self.focus_force()

            except Exception:
                pass

    # =========================================================
    # EDITAR UBICACIÓN
    # =========================================================

    def editar_ubicacion(self):

        seleccionado = self.tabla.selection()

        # =================================================
        # VALIDAR SELECCIÓN
        # =================================================

        if not seleccionado:

            messagebox.showwarning(
                "Advertencia",
                (
                    "Seleccione una ubicación "
                    "para editar."
                ),
                parent=self
            )

            return

        # =================================================
        # COMPROBAR FORMULARIO ABIERTO
        # =================================================

        if (
            self.ventana_formulario is not None
            and self.ventana_formulario.winfo_exists()
        ):

            self.ventana_formulario.lift()
            self.ventana_formulario.focus_force()

            return

        # =================================================
        # OBTENER DATOS
        # =================================================

        datos = self.tabla.item(
            seleccionado[0],
            "values"
        )

        if not datos:

            messagebox.showerror(
                "Error",
                "No se pudieron obtener los datos de la ubicación.",
                parent=self
            )

            return

        try:

            from app.views.formulario_ubicacion import (
                FormularioUbicacion
            )

            # =================================================
            # ABRIR FORMULARIO
            # =================================================

            self.ventana_formulario = FormularioUbicacion(
                self,
                ubicacion=datos,
                callback=self.cargar_ubicaciones
            )

            self.ventana_formulario.transient(
                self
            )

            self.ventana_formulario.lift()
            self.ventana_formulario.focus_force()

            self.ventana_formulario.grab_set()

            self.ventana_formulario.protocol(
                "WM_DELETE_WINDOW",
                self.cerrar_formulario
            )

        except Exception as e:

            self.ventana_formulario = None

            print(
                "Error al editar ubicación:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudo abrir el formulario "
                    "de edición:\n\n"
                    f"{e}"
                ),
                parent=self
            )

    # =========================================================
    # ELIMINAR UBICACIÓN
    # =========================================================

    def eliminar_ubicacion(self):

        seleccionado = self.tabla.selection()

        # =================================================
        # VALIDAR SELECCIÓN
        # =================================================

        if not seleccionado:

            messagebox.showwarning(
                "Advertencia",
                (
                    "Seleccione una ubicación "
                    "para eliminar."
                ),
                parent=self
            )

            return

        # =================================================
        # OBTENER DATOS
        # =================================================

        datos = self.tabla.item(
            seleccionado[0],
            "values"
        )

        if not datos:

            messagebox.showerror(
                "Error",
                "No se pudieron obtener los datos de la ubicación.",
                parent=self
            )

            return

        id_ubicacion = datos[0]
        nombre = datos[1]

        # =================================================
        # CONFIRMAR
        # =================================================

        confirmar = messagebox.askyesno(
            "Confirmar eliminación",
            (
                "¿Está seguro de eliminar "
                f"la ubicación '{nombre}'?"
            ),
            parent=self
        )

        if not confirmar:
            return

        try:

            # =================================================
            # ELIMINAR
            # =================================================

            self.model.eliminar(
                id_ubicacion
            )

            # =================================================
            # MENSAJE
            # =================================================

            messagebox.showinfo(
                "Éxito",
                (
                    "Ubicación eliminada "
                    "correctamente."
                ),
                parent=self
            )

            # =================================================
            # RECARGAR
            # =================================================

            self.cargar_ubicaciones()

        except Exception as e:

            print(
                "Error al eliminar ubicación:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudo eliminar "
                    "la ubicación.\n\n"
                    f"{e}"
                ),
                parent=self
            )