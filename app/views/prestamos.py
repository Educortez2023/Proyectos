import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.models.prestamo_model import PrestamoModel


class VentanaPrestamos(ctk.CTkToplevel):

    def __init__(self, parent):

        # =====================================================
        # CREAR VENTANA
        # =====================================================

        super().__init__(parent)

        self.parent = parent

        # =====================================================
        # CONFIGURACIÓN GENERAL
        # =====================================================

        self.title(
            "Gestión de Préstamos y Asignaciones"
        )

        self.transient(parent)

        self.resizable(
            True,
            True
        )

        self.minsize(
            950,
            550
        )

        # =====================================================
        # OCULTAR VENTANA DURANTE LA CONSTRUCCIÓN
        # =====================================================

        self.withdraw()

        # =====================================================
        # MODELO
        # =====================================================

        self.model = PrestamoModel()

        # =====================================================
        # REFERENCIA AL FORMULARIO
        # =====================================================

        self.ventana_formulario = None

        # =====================================================
        # CREAR INTERFAZ
        # =====================================================

        self.crear_interfaz()

        # =====================================================
        # CONFIGURAR TAMAÑO
        # =====================================================

        self.configurar_tamano_inicial()

        # =====================================================
        # CARGAR DATOS
        # =====================================================

        self.cargar_prestamos()

        # =====================================================
        # ACTUALIZAR GEOMETRÍA INTERNA
        # =====================================================

        self.update_idletasks()

        # =====================================================
        # AJUSTAR COLUMNAS UNA SOLA VEZ
        # =====================================================

        self.ajustar_columnas()

        # =====================================================
        # MOSTRAR VENTANA
        # =====================================================

        self.deiconify()

        self.lift()

        self.focus_force()

    # =========================================================
    # CONFIGURAR TAMAÑO INICIAL
    # =========================================================

    def configurar_tamano_inicial(self):

        try:

            ancho_pantalla = self.winfo_screenwidth()
            alto_pantalla = self.winfo_screenheight()

            # -------------------------------------------------
            # TAMAÑO DESEADO
            # -------------------------------------------------

            ancho = 1250
            alto = 650

            # -------------------------------------------------
            # NO SUPERAR LA PANTALLA
            # -------------------------------------------------

            ancho = min(
                ancho,
                ancho_pantalla - 80
            )

            alto = min(
                alto,
                alto_pantalla - 100
            )

            # -------------------------------------------------
            # RESPETAR MÍNIMOS
            # -------------------------------------------------

            ancho = max(
                ancho,
                950
            )

            alto = max(
                alto,
                550
            )

            # -------------------------------------------------
            # CENTRAR
            # -------------------------------------------------

            x = (
                ancho_pantalla - ancho
            ) // 2

            y = (
                alto_pantalla - alto
            ) // 2

            self.geometry(
                f"{ancho}x{alto}+{x}+{y}"
            )

        except Exception as e:

            print(
                "Error al configurar tamaño:",
                e
            )

            self.geometry(
                "1250x650"
            )

    # =========================================================
    # AJUSTAR COLUMNAS
    # =========================================================

    def ajustar_columnas(self):

        try:

            if not hasattr(
                self,
                "tabla"
            ):

                return

            self.update_idletasks()

            # -------------------------------------------------
            # ANCHO DISPONIBLE
            # -------------------------------------------------

            ancho_disponible = (
                self.frame_tabla.winfo_width()
            )

            if ancho_disponible <= 1:

                return

            # Reservar scrollbar vertical
            ancho_disponible -= 20

            # -------------------------------------------------
            # ANCHOS MÍNIMOS
            # -------------------------------------------------

            minimos = {
                "id": 55,
                "codigo": 95,
                "equipo": 145,
                "responsable": 155,
                "fecha_prestamo": 115,
                "fecha_devolucion": 120,
                "estado": 90,
                "observaciones": 180
            }

            columnas = (
                "id",
                "codigo",
                "equipo",
                "responsable",
                "fecha_prestamo",
                "fecha_devolucion",
                "estado",
                "observaciones"
            )

            # -------------------------------------------------
            # PESOS
            # -------------------------------------------------

            pesos = {
                "id": 0.04,
                "codigo": 0.07,
                "equipo": 0.16,
                "responsable": 0.18,
                "fecha_prestamo": 0.11,
                "fecha_devolucion": 0.12,
                "estado": 0.09,
                "observaciones": 0.23
            }

            # -------------------------------------------------
            # SUMA DE MÍNIMOS
            # -------------------------------------------------

            suma_minimos = sum(
                minimos[columna]
                for columna in columnas
            )

            # =================================================
            # HAY SUFICIENTE ESPACIO
            # =================================================

            if ancho_disponible >= suma_minimos:

                espacio_extra = (
                    ancho_disponible
                    - suma_minimos
                )

                for columna in columnas:

                    ancho = (
                        minimos[columna]
                        + (
                            espacio_extra
                            * pesos[columna]
                        )
                    )

                    self.tabla.column(
                        columna,
                        width=int(ancho),
                        minwidth=minimos[columna],
                        stretch=False
                    )

            # =================================================
            # NO HAY SUFICIENTE ESPACIO
            # =================================================

            else:

                for columna in columnas:

                    self.tabla.column(
                        columna,
                        width=minimos[columna],
                        minwidth=minimos[columna],
                        stretch=False
                    )

        except Exception as e:

            print(
                "Error al ajustar columnas:",
                e
            )

    # =========================================================
    # CREAR INTERFAZ
    # =========================================================

    def crear_interfaz(self):

        # =====================================================
        # CONTENEDOR PRINCIPAL
        # =====================================================

        frame_principal = ctk.CTkFrame(
            self,
            fg_color="transparent"
        )

        frame_principal.pack(
            fill="both",
            expand=True,
            padx=15,
            pady=10
        )

        self.frame_principal = frame_principal

        # =====================================================
        # TÍTULO
        # =====================================================

        titulo = ctk.CTkLabel(
            frame_principal,
            text="Gestión de Préstamos y Asignaciones",
            font=("Arial", 24, "bold")
        )

        titulo.pack(
            fill="x",
            pady=(5, 10)
        )

        # =====================================================
        # CONTROLES
        # =====================================================

        frame_controles = ctk.CTkFrame(
            frame_principal
        )

        frame_controles.pack(
            fill="x",
            pady=(0, 10)
        )

        # =====================================================
        # FRAME BOTONES
        # =====================================================

        frame_botones = ctk.CTkFrame(
            frame_controles,
            fg_color="transparent"
        )

        frame_botones.pack(
            side="left",
            padx=8,
            pady=8
        )

        self.frame_botones = frame_botones

        # =====================================================
        # BOTÓN NUEVO
        # =====================================================

        btn_nuevo = ctk.CTkButton(
            frame_botones,
            text="➕ Nuevo Préstamo",
            width=145,
            height=36,
            command=self.abrir_nuevo
        )

        btn_nuevo.grid(
            row=0,
            column=0,
            padx=3
        )

        # =====================================================
        # BOTÓN EDITAR
        # =====================================================

        btn_editar = ctk.CTkButton(
            frame_botones,
            text="✏ Editar",
            width=100,
            height=36,
            command=self.editar_prestamo
        )

        btn_editar.grid(
            row=0,
            column=1,
            padx=3
        )

        # =====================================================
        # BOTÓN DEVOLUCIÓN
        # =====================================================

        btn_devolucion = ctk.CTkButton(
            frame_botones,
            text="↩ Devolución",
            width=115,
            height=36,
            command=self.registrar_devolucion
        )

        btn_devolucion.grid(
            row=0,
            column=2,
            padx=3
        )

        # =====================================================
        # BOTÓN ELIMINAR
        # =====================================================

        btn_eliminar = ctk.CTkButton(
            frame_botones,
            text="🗑 Eliminar",
            width=105,
            height=36,
            command=self.eliminar_prestamo
        )

        btn_eliminar.grid(
            row=0,
            column=3,
            padx=3
        )

        # =====================================================
        # BOTÓN ACTUALIZAR
        # =====================================================

        btn_actualizar = ctk.CTkButton(
            frame_botones,
            text="🔄 Actualizar",
            width=110,
            height=36,
            command=self.cargar_prestamos
        )

        btn_actualizar.grid(
            row=0,
            column=4,
            padx=3
        )

        # =====================================================
        # BOTÓN EXCEL
        # =====================================================

        btn_exportar = ctk.CTkButton(
            frame_botones,
            text="📊 Excel",
            width=100,
            height=36,
            command=self.exportar_excel
        )

        btn_exportar.grid(
            row=0,
            column=5,
            padx=3
        )

        # =====================================================
        # FRAME BÚSQUEDA
        # =====================================================

        frame_busqueda = ctk.CTkFrame(
            frame_controles,
            fg_color="transparent"
        )

        frame_busqueda.pack(
            side="right",
            padx=8,
            pady=8
        )

        self.frame_busqueda = frame_busqueda

        # =====================================================
        # CAMPO BÚSQUEDA
        # =====================================================

        self.entry_busqueda = ctk.CTkEntry(
            frame_busqueda,
            width=280,
            height=36,
            placeholder_text="Buscar préstamo..."
        )

        self.entry_busqueda.grid(
            row=0,
            column=0,
            padx=3
        )

        # =====================================================
        # BOTÓN BUSCAR
        # =====================================================

        btn_buscar = ctk.CTkButton(
            frame_busqueda,
            text="🔍 Buscar",
            width=100,
            height=36,
            command=self.buscar_prestamos
        )

        btn_buscar.grid(
            row=0,
            column=1,
            padx=3
        )

        # =====================================================
        # ENTER PARA BUSCAR
        # =====================================================

        self.entry_busqueda.bind(
            "<Return>",
            lambda event: self.buscar_prestamos()
        )

        # =====================================================
        # FRAME TABLA
        # =====================================================

        self.frame_tabla = ctk.CTkFrame(
            frame_principal
        )

        self.frame_tabla.pack(
            fill="both",
            expand=True
        )

        # =====================================================
        # GRID TABLA
        # =====================================================

        self.frame_tabla.grid_rowconfigure(
            0,
            weight=1
        )

        self.frame_tabla.grid_columnconfigure(
            0,
            weight=1
        )

        # =====================================================
        # SCROLL VERTICAL
        # =====================================================

        scrollbar_vertical = ttk.Scrollbar(
            self.frame_tabla,
            orient="vertical"
        )

        scrollbar_vertical.grid(
            row=0,
            column=1,
            sticky="ns"
        )

        # =====================================================
        # SCROLL HORIZONTAL
        # =====================================================

        scrollbar_horizontal = ttk.Scrollbar(
            self.frame_tabla,
            orient="horizontal"
        )

        scrollbar_horizontal.grid(
            row=1,
            column=0,
            sticky="ew"
        )

        # =====================================================
        # COLUMNAS
        # =====================================================

        columnas = (
            "id",
            "codigo",
            "equipo",
            "responsable",
            "fecha_prestamo",
            "fecha_devolucion",
            "estado",
            "observaciones"
        )

        # =====================================================
        # TREEVIEW
        # =====================================================

        self.tabla = ttk.Treeview(
            self.frame_tabla,
            columns=columnas,
            show="headings",
            selectmode="browse",
            yscrollcommand=scrollbar_vertical.set,
            xscrollcommand=scrollbar_horizontal.set
        )

        self.tabla.grid(
            row=0,
            column=0,
            sticky="nsew"
        )

        # =====================================================
        # CONECTAR SCROLL VERTICAL
        # =====================================================

        scrollbar_vertical.configure(
            command=self.tabla.yview
        )

        # =====================================================
        # CONECTAR SCROLL HORIZONTAL
        # =====================================================

        scrollbar_horizontal.configure(
            command=self.tabla.xview
        )

        # =====================================================
        # ENCABEZADOS
        # =====================================================

        self.tabla.heading(
            "id",
            text="ID"
        )

        self.tabla.heading(
            "codigo",
            text="Código"
        )

        self.tabla.heading(
            "equipo",
            text="Equipo"
        )

        self.tabla.heading(
            "responsable",
            text="Responsable"
        )

        self.tabla.heading(
            "fecha_prestamo",
            text="Fecha Préstamo"
        )

        self.tabla.heading(
            "fecha_devolucion",
            text="Fecha Devolución"
        )

        self.tabla.heading(
            "estado",
            text="Estado"
        )

        self.tabla.heading(
            "observaciones",
            text="Observaciones"
        )

        # =====================================================
        # ANCHOS
        # =====================================================

        self.tabla.column(
            "id",
            width=55,
            minwidth=55,
            anchor="center",
            stretch=False
        )

        self.tabla.column(
            "codigo",
            width=95,
            minwidth=95,
            anchor="center",
            stretch=False
        )

        self.tabla.column(
            "equipo",
            width=145,
            minwidth=145,
            anchor="w",
            stretch=False
        )

        self.tabla.column(
            "responsable",
            width=155,
            minwidth=155,
            anchor="w",
            stretch=False
        )

        self.tabla.column(
            "fecha_prestamo",
            width=115,
            minwidth=115,
            anchor="center",
            stretch=False
        )

        self.tabla.column(
            "fecha_devolucion",
            width=120,
            minwidth=120,
            anchor="center",
            stretch=False
        )

        self.tabla.column(
            "estado",
            width=90,
            minwidth=90,
            anchor="center",
            stretch=False
        )

        self.tabla.column(
            "observaciones",
            width=180,
            minwidth=180,
            anchor="w",
            stretch=False
        )

        # =====================================================
        # DOBLE CLICK
        # =====================================================

        self.tabla.bind(
            "<Double-1>",
            lambda event: self.editar_prestamo()
        )

    # =========================================================
    # CARGAR PRÉSTAMOS
    # =========================================================

    def cargar_prestamos(self):

        try:

            for item in self.tabla.get_children():

                self.tabla.delete(
                    item
                )

            prestamos = self.model.listar()

            for prestamo in prestamos:

                self.tabla.insert(
                    "",
                    "end",
                    values=(
                        prestamo[0],
                        prestamo[2],
                        prestamo[3],
                        prestamo[5],
                        prestamo[6],
                        prestamo[7],
                        prestamo[8],
                        prestamo[9]
                    )
                )

            print(
                "Préstamos cargados:",
                len(prestamos)
            )

        except Exception as e:

            print(
                "Error al cargar préstamos:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudieron cargar "
                    "los préstamos:\n\n"
                    f"{e}"
                ),
                parent=self
            )

    # =========================================================
    # BUSCAR PRÉSTAMOS
    # =========================================================

    def buscar_prestamos(self):

        try:

            texto = (
                self.entry_busqueda
                .get()
                .strip()
            )

            if not texto:

                self.cargar_prestamos()

                return

            for item in self.tabla.get_children():

                self.tabla.delete(
                    item
                )

            prestamos = self.model.buscar(
                texto
            )

            for prestamo in prestamos:

                self.tabla.insert(
                    "",
                    "end",
                    values=(
                        prestamo[0],
                        prestamo[2],
                        prestamo[3],
                        prestamo[5],
                        prestamo[6],
                        prestamo[7],
                        prestamo[8],
                        prestamo[9]
                    )
                )

            print(
                "Préstamos encontrados:",
                len(prestamos)
            )

        except Exception as e:

            print(
                "Error en búsqueda:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudo realizar "
                    "la búsqueda:\n\n"
                    f"{e}"
                ),
                parent=self
            )

    # =========================================================
    # REGISTRAR DEVOLUCIÓN
    # =========================================================

    def registrar_devolucion(self):

        seleccionado = self.tabla.selection()

        if not seleccionado:

            messagebox.showwarning(
                "Advertencia",
                "Seleccione un préstamo para registrar la devolución.",
                parent=self
            )

            return

        datos = self.tabla.item(
            seleccionado[0],
            "values"
        )

        if not datos:

            messagebox.showerror(
                "Error",
                "No se pudieron obtener los datos del préstamo.",
                parent=self
            )

            return

        id_prestamo = datos[0]
        codigo = datos[1]
        equipo = datos[2]
        responsable = datos[3]
        estado = datos[6]

        if estado == "DEVUELTO":

            messagebox.showinfo(
                "Préstamo ya devuelto",
                (
                    "Este préstamo ya se encuentra "
                    "registrado como DEVUELTO."
                ),
                parent=self
            )

            return

        confirmar = messagebox.askyesno(
            "Confirmar devolución",
            (
                "¿Desea registrar la devolución de este equipo?\n\n"
                f"Código: {codigo}\n"
                f"Equipo: {equipo}\n"
                f"Responsable: {responsable}\n\n"
                "La fecha de devolución será registrada "
                "automáticamente con la fecha actual."
            ),
            parent=self
        )

        if not confirmar:

            return

        try:

            self.model.registrar_devolucion(
                id_prestamo
            )

            messagebox.showinfo(
                "Devolución registrada",
                (
                    "La devolución se registró "
                    "correctamente.\n\n"
                    f"Equipo: {equipo}\n"
                    "Estado: DEVUELTO"
                ),
                parent=self
            )

            self.cargar_prestamos()

        except Exception as e:

            print(
                "Error al registrar devolución:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudo registrar "
                    "la devolución:\n\n"
                    f"{e}"
                ),
                parent=self
            )

    # =========================================================
    # EXPORTAR A EXCEL
    # =========================================================

    def exportar_excel(self):

        try:

            prestamos = self.model.listar()

            if not prestamos:

                messagebox.showwarning(
                    "Sin datos",
                    (
                        "No existen préstamos "
                        "registrados para exportar."
                    ),
                    parent=self
                )

                return

            ruta = filedialog.asksaveasfilename(
                parent=self,
                title="Guardar reporte de préstamos",
                defaultextension=".xlsx",
                filetypes=[
                    (
                        "Archivo Excel",
                        "*.xlsx"
                    ),
                    (
                        "Todos los archivos",
                        "*.*"
                    )
                ],
                initialfile="Reporte_Prestamos.xlsx"
            )

            if not ruta:

                return

            libro = Workbook()

            hoja = libro.active

            hoja.title = "Préstamos"

            encabezados = [
                "ID",
                "Código",
                "Equipo",
                "Responsable",
                "Fecha Préstamo",
                "Fecha Devolución",
                "Estado",
                "Observaciones"
            ]

            hoja.append(
                encabezados
            )

            for celda in hoja[1]:

                celda.font = Font(
                    bold=True
                )

            for prestamo in prestamos:

                hoja.append(
                    [
                        prestamo[0],
                        prestamo[2],
                        prestamo[3],
                        prestamo[5],
                        prestamo[6],
                        prestamo[7],
                        prestamo[8],
                        prestamo[9]
                    ]
                )

            ultima_fila = hoja.max_row

            ultima_columna = hoja.max_column

            ultima_columna_letra = get_column_letter(
                ultima_columna
            )

            hoja.auto_filter.ref = (
                f"A1:{ultima_columna_letra}{ultima_fila}"
            )

            hoja.freeze_panes = "A2"

            for columna in hoja.columns:

                longitud_maxima = 0

                numero_columna = (
                    columna[0].column
                )

                letra_columna = get_column_letter(
                    numero_columna
                )

                for celda in columna:

                    if celda.value is not None:

                        longitud = len(
                            str(
                                celda.value
                            )
                        )

                        if longitud > longitud_maxima:

                            longitud_maxima = longitud

                hoja.column_dimensions[
                    letra_columna
                ].width = min(
                    longitud_maxima + 2,
                    50
                )

            libro.save(
                ruta
            )

            messagebox.showinfo(
                "Exportación exitosa",
                (
                    "El reporte de préstamos "
                    "se exportó correctamente "
                    "a Excel.\n\n"
                    f"Registros exportados: "
                    f"{len(prestamos)}"
                ),
                parent=self
            )

            print(
                "Excel de préstamos exportado:",
                ruta
            )

        except Exception as e:

            print(
                "Error al exportar préstamos:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudo exportar "
                    "el archivo:\n\n"
                    f"{e}"
                ),
                parent=self
            )

    # =========================================================
    # ABRIR NUEVO PRÉSTAMO
    # =========================================================

    def abrir_nuevo(self):

        if (
            self.ventana_formulario is not None
            and self.ventana_formulario.winfo_exists()
        ):

            self.ventana_formulario.lift()

            self.ventana_formulario.focus_force()

            return

        try:

            from app.views.formulario_prestamo import (
                FormularioPrestamo
            )

            self.ventana_formulario = FormularioPrestamo(
                self,
                callback=self.cargar_prestamos
            )

            self.ventana_formulario.transient(
                self
            )

            self.ventana_formulario.lift()

            self.ventana_formulario.focus_force()

            self.ventana_formulario.grab_set()

            self.ventana_formulario.protocol(
                "WM_DELETE_WINDOW",
                self.cerrar_formulario
            )

        except Exception as e:

            self.ventana_formulario = None

            print(
                "Error al abrir formulario:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudo abrir el formulario "
                    "de préstamo:\n\n"
                    f"{e}"
                ),
                parent=self
            )

    # =========================================================
    # CERRAR FORMULARIO
    # =========================================================

    def cerrar_formulario(self):

        try:

            if (
                self.ventana_formulario is not None
                and self.ventana_formulario.winfo_exists()
            ):

                self.ventana_formulario.grab_release()

                self.ventana_formulario.destroy()

        except Exception as e:

            print(
                "Error al cerrar formulario:",
                e
            )

        finally:

            self.ventana_formulario = None

            try:

                self.lift()

                self.focus_force()

            except Exception:
                pass

    # =========================================================
    # EDITAR PRÉSTAMO
    # =========================================================

    def editar_prestamo(self):

        seleccionado = self.tabla.selection()

        if not seleccionado:

            messagebox.showwarning(
                "Advertencia",
                "Seleccione un préstamo para editar.",
                parent=self
            )

            return

        if (
            self.ventana_formulario is not None
            and self.ventana_formulario.winfo_exists()
        ):

            self.ventana_formulario.lift()

            self.ventana_formulario.focus_force()

            return

        datos = self.tabla.item(
            seleccionado[0],
            "values"
        )

        if not datos:

            messagebox.showerror(
                "Error",
                "No se pudieron obtener los datos del préstamo.",
                parent=self
            )

            return

        try:

            from app.views.formulario_prestamo import (
                FormularioPrestamo
            )

            self.ventana_formulario = FormularioPrestamo(
                self,
                prestamo=datos,
                callback=self.cargar_prestamos
            )

            self.ventana_formulario.transient(
                self
            )

            self.ventana_formulario.lift()

            self.ventana_formulario.focus_force()

            self.ventana_formulario.grab_set()

            self.ventana_formulario.protocol(
                "WM_DELETE_WINDOW",
                self.cerrar_formulario
            )

        except Exception as e:

            self.ventana_formulario = None

            print(
                "Error al editar préstamo:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudo abrir el formulario "
                    "de edición:\n\n"
                    f"{e}"
                ),
                parent=self
            )

    # =========================================================
    # ELIMINAR PRÉSTAMO
    # =========================================================

    def eliminar_prestamo(self):

        seleccionado = self.tabla.selection()

        if not seleccionado:

            messagebox.showwarning(
                "Advertencia",
                "Seleccione un préstamo para eliminar.",
                parent=self
            )

            return

        datos = self.tabla.item(
            seleccionado[0],
            "values"
        )

        if not datos:

            messagebox.showerror(
                "Error",
                "No se pudieron obtener los datos del préstamo.",
                parent=self
            )

            return

        id_prestamo = datos[0]

        confirmar = messagebox.askyesno(
            "Confirmar eliminación",
            (
                "¿Está seguro de eliminar "
                "este préstamo?"
            ),
            parent=self
        )

        if not confirmar:

            return

        try:

            self.model.eliminar(
                id_prestamo
            )

            messagebox.showinfo(
                "Éxito",
                "Préstamo eliminado correctamente.",
                parent=self
            )

            self.cargar_prestamos()

        except Exception as e:

            print(
                "Error al eliminar préstamo:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudo eliminar "
                    "el préstamo:\n\n"
                    f"{e}"
                ),
                parent=self
            )