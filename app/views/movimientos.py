import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.models.movimiento_model import MovimientoModel


class VentanaMovimientos(ctk.CTkToplevel):

    def __init__(self, parent):

        super().__init__(parent)

        # =====================================================
        # OCULTAR VENTANA DURANTE LA CONFIGURACIÓN
        # =====================================================

        self.withdraw()

        # =====================================================
        # CONFIGURACIÓN DE VENTANA
        # =====================================================

        self.parent = parent

        self.title(
            "Gestión de Movimientos de Insumos"
        )

        self.minsize(
            1150,
            600
        )

        self.transient(parent)

        # =====================================================
        # MODELO
        # =====================================================

        self.model = MovimientoModel()

        # =====================================================
        # REFERENCIA AL FORMULARIO
        # =====================================================

        self.ventana_formulario = None

        # =====================================================
        # CREAR INTERFAZ
        # =====================================================

        self.crear_interfaz()

        # =====================================================
        # CARGAR DATOS
        # =====================================================

        self.cargar_movimientos()

        # =====================================================
        # MOSTRAR MAXIMIZADA
        # =====================================================

        self.after(
            50,
            self.mostrar_delante
        )


    # =====================================================
    # MOSTRAR VENTANA MAXIMIZADA
    # =====================================================

    def mostrar_delante(self):

        try:

            # ---------------------------------------------
            # MAXIMIZAR
            # ---------------------------------------------

            self.state(
                "zoomed"
            )

        except Exception as e:

            print(
                "No se pudo maximizar con state:",
                e
            )

            # ---------------------------------------------
            # RESPALDO
            # ---------------------------------------------

            try:

                ancho = self.winfo_screenwidth()
                alto = self.winfo_screenheight()

                self.geometry(
                    f"{ancho}x{alto}+0+0"
                )

            except Exception as error:

                print(
                    "No se pudo ajustar la ventana:",
                    error
                )

        # ---------------------------------------------
        # MOSTRAR VENTANA
        # ---------------------------------------------

        self.deiconify()

        self.lift()

        self.focus_force()


    # =====================================================
    # CREAR INTERFAZ
    # =====================================================

    def crear_interfaz(self):

        # =================================================
        # TÍTULO
        # =================================================

        titulo = ctk.CTkLabel(
            self,
            text="Gestión de Movimientos de Insumos",
            font=("Arial", 24, "bold")
        )

        titulo.pack(
            pady=(20, 10)
        )

        # =================================================
        # FRAME DE CONTROLES
        # =================================================

        frame_controles = ctk.CTkFrame(
            self
        )

        frame_controles.pack(
            fill="x",
            padx=20,
            pady=10
        )

        # =================================================
        # BOTÓN NUEVO
        # =================================================

        btn_nuevo = ctk.CTkButton(
            frame_controles,
            text="➕ Nuevo Movimiento",
            width=170,
            command=self.abrir_nuevo
        )

        btn_nuevo.pack(
            side="left",
            padx=5,
            pady=10
        )

        # =================================================
        # BOTÓN ACTUALIZAR
        # =================================================

        btn_actualizar = ctk.CTkButton(
            frame_controles,
            text="🔄 Actualizar",
            width=120,
            command=self.cargar_movimientos
        )

        btn_actualizar.pack(
            side="left",
            padx=5,
            pady=10
        )

        # =================================================
        # BOTÓN EXPORTAR
        # =================================================

        btn_exportar = ctk.CTkButton(
            frame_controles,
            text="📊 Exportar a Excel",
            width=160,
            command=self.exportar_excel
        )

        btn_exportar.pack(
            side="left",
            padx=5,
            pady=10
        )

        # =================================================
        # CAMPO DE BÚSQUEDA
        # =================================================

        self.entry_busqueda = ctk.CTkEntry(
            frame_controles,
            width=300,
            placeholder_text="Buscar movimiento..."
        )

        self.entry_busqueda.pack(
            side="right",
            padx=5,
            pady=10
        )

        # =================================================
        # BOTÓN BUSCAR
        # =================================================

        btn_buscar = ctk.CTkButton(
            frame_controles,
            text="🔍 Buscar",
            width=110,
            command=self.buscar_movimientos
        )

        btn_buscar.pack(
            side="right",
            padx=5,
            pady=10
        )

        # =================================================
        # ENTER PARA BUSCAR
        # =================================================

        self.entry_busqueda.bind(
            "<Return>",
            lambda event: self.buscar_movimientos()
        )

        # =================================================
        # FRAME TABLA
        # =================================================

        frame_tabla = ctk.CTkFrame(
            self
        )

        frame_tabla.pack(
            fill="both",
            expand=True,
            padx=20,
            pady=(5, 20)
        )

        # =================================================
        # SCROLLBAR VERTICAL
        # =================================================

        scrollbar_vertical = ttk.Scrollbar(
            frame_tabla,
            orient="vertical"
        )

        scrollbar_vertical.pack(
            side="right",
            fill="y"
        )

        # =================================================
        # SCROLLBAR HORIZONTAL
        # =================================================

        scrollbar_horizontal = ttk.Scrollbar(
            frame_tabla,
            orient="horizontal"
        )

        scrollbar_horizontal.pack(
            side="bottom",
            fill="x"
        )

        # =================================================
        # COLUMNAS
        # =================================================

        columnas = (
            "id",
            "codigo",
            "insumo",
            "tipo",
            "cantidad",
            "stock_anterior",
            "stock_nuevo",
            "responsable",
            "observaciones",
            "fecha"
        )

        # =================================================
        # CREAR TABLA
        # =================================================

        self.tabla = ttk.Treeview(
            frame_tabla,
            columns=columnas,
            show="headings",
            yscrollcommand=scrollbar_vertical.set,
            xscrollcommand=scrollbar_horizontal.set
        )

        # =================================================
        # ENCABEZADOS
        # =================================================

        self.tabla.heading(
            "id",
            text="ID"
        )

        self.tabla.heading(
            "codigo",
            text="Código"
        )

        self.tabla.heading(
            "insumo",
            text="Insumo"
        )

        self.tabla.heading(
            "tipo",
            text="Movimiento"
        )

        self.tabla.heading(
            "cantidad",
            text="Cantidad"
        )

        self.tabla.heading(
            "stock_anterior",
            text="Stock Anterior"
        )

        self.tabla.heading(
            "stock_nuevo",
            text="Stock Nuevo"
        )

        self.tabla.heading(
            "responsable",
            text="Responsable"
        )

        self.tabla.heading(
            "observaciones",
            text="Observaciones"
        )

        self.tabla.heading(
            "fecha",
            text="Fecha"
        )

        # =================================================
        # ANCHOS
        # =================================================

        self.tabla.column(
            "id",
            width=60,
            anchor="center"
        )

        self.tabla.column(
            "codigo",
            width=120,
            anchor="center"
        )

        self.tabla.column(
            "insumo",
            width=220
        )

        self.tabla.column(
            "tipo",
            width=120,
            anchor="center"
        )

        self.tabla.column(
            "cantidad",
            width=90,
            anchor="center"
        )

        self.tabla.column(
            "stock_anterior",
            width=120,
            anchor="center"
        )

        self.tabla.column(
            "stock_nuevo",
            width=110,
            anchor="center"
        )

        self.tabla.column(
            "responsable",
            width=220
        )

        self.tabla.column(
            "observaciones",
            width=260
        )

        self.tabla.column(
            "fecha",
            width=160,
            anchor="center"
        )

        # =================================================
        # MOSTRAR TABLA
        # =================================================

        self.tabla.pack(
            fill="both",
            expand=True
        )

        # =================================================
        # CONECTAR SCROLLBARS
        # =================================================

        scrollbar_vertical.config(
            command=self.tabla.yview
        )

        scrollbar_horizontal.config(
            command=self.tabla.xview
        )


    # =====================================================
    # CARGAR MOVIMIENTOS
    # =====================================================

    def cargar_movimientos(self):

        try:

            # =================================================
            # LIMPIAR TABLA
            # =================================================

            for item in self.tabla.get_children():

                self.tabla.delete(
                    item
                )

            # =================================================
            # OBTENER DATOS DESDE LA BD
            # =================================================

            movimientos = self.model.listar()

            # =================================================
            # INSERTAR DATOS
            # =================================================

            for movimiento in movimientos:

                self.tabla.insert(
                    "",
                    "end",
                    values=(
                        movimiento[0],
                        movimiento[2],
                        movimiento[3],
                        movimiento[5],
                        movimiento[6],
                        movimiento[7],
                        movimiento[8],
                        movimiento[10],
                        movimiento[11],
                        movimiento[12]
                    )
                )

            print(
                "Movimientos cargados:",
                len(movimientos)
            )

        except Exception as e:

            print(
                "Error al cargar movimientos:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudieron cargar "
                    "los movimientos:\n\n"
                    f"{e}"
                ),
                parent=self
            )


    # =====================================================
    # BUSCAR MOVIMIENTOS
    # =====================================================

    def buscar_movimientos(self):

        try:

            texto = (
                self.entry_busqueda
                .get()
                .strip()
            )

            # =================================================
            # SI ESTÁ VACÍO
            # =================================================

            if not texto:

                self.cargar_movimientos()

                return

            # =================================================
            # LIMPIAR TABLA
            # =================================================

            for item in self.tabla.get_children():

                self.tabla.delete(
                    item
                )

            # =================================================
            # BUSCAR EN LA BD
            # =================================================

            movimientos = self.model.buscar(
                texto
            )

            # =================================================
            # MOSTRAR RESULTADOS
            # =================================================

            for movimiento in movimientos:

                self.tabla.insert(
                    "",
                    "end",
                    values=(
                        movimiento[0],
                        movimiento[2],
                        movimiento[3],
                        movimiento[5],
                        movimiento[6],
                        movimiento[7],
                        movimiento[8],
                        movimiento[10],
                        movimiento[11],
                        movimiento[12]
                    )
                )

            print(
                "Movimientos encontrados:",
                len(movimientos)
            )

        except Exception as e:

            print(
                "Error en búsqueda:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudo realizar "
                    "la búsqueda:\n\n"
                    f"{e}"
                ),
                parent=self
            )


    # =====================================================
    # EXPORTAR A EXCEL
    # =====================================================

    def exportar_excel(self):

        try:

            # =================================================
            # OBTENER DATOS DESDE LA BD
            # =================================================

            movimientos = self.model.listar()

            # =================================================
            # VALIDAR
            # =================================================

            if not movimientos:

                messagebox.showwarning(
                    "Sin datos",
                    (
                        "No existen movimientos "
                        "registrados para exportar."
                    ),
                    parent=self
                )

                return

            # =================================================
            # SELECCIONAR ARCHIVO
            # =================================================

            ruta = filedialog.asksaveasfilename(
                parent=self,
                title="Guardar reporte de movimientos",
                defaultextension=".xlsx",
                filetypes=[
                    (
                        "Archivo Excel",
                        "*.xlsx"
                    ),
                    (
                        "Todos los archivos",
                        "*.*"
                    )
                ],
                initialfile="Reporte_Movimientos.xlsx"
            )

            if not ruta:

                return

            # =================================================
            # CREAR LIBRO
            # =================================================

            libro = Workbook()

            hoja = libro.active

            hoja.title = "Movimientos"

            # =================================================
            # ENCABEZADOS
            # =================================================

            encabezados = [
                "ID",
                "Código",
                "Insumo",
                "Movimiento",
                "Cantidad",
                "Stock Anterior",
                "Stock Nuevo",
                "Responsable",
                "Observaciones",
                "Fecha"
            ]

            hoja.append(
                encabezados
            )

            # =================================================
            # FORMATO
            # =================================================

            for celda in hoja[1]:

                celda.font = Font(
                    bold=True
                )

            # =================================================
            # INSERTAR DATOS
            # =================================================

            for movimiento in movimientos:

                hoja.append(
                    [
                        movimiento[0],
                        movimiento[2],
                        movimiento[3],
                        movimiento[5],
                        movimiento[6],
                        movimiento[7],
                        movimiento[8],
                        movimiento[10],
                        movimiento[11],
                        movimiento[12]
                    ]
                )

            # =================================================
            # FILTRO AUTOMÁTICO
            # =================================================

            ultima_fila = hoja.max_row

            ultima_columna = hoja.max_column

            ultima_columna_letra = get_column_letter(
                ultima_columna
            )

            hoja.auto_filter.ref = (
                f"A1:{ultima_columna_letra}{ultima_fila}"
            )

            # =================================================
            # CONGELAR ENCABEZADOS
            # =================================================

            hoja.freeze_panes = "A2"

            # =================================================
            # AJUSTAR COLUMNAS
            # =================================================

            for columna in hoja.columns:

                longitud_maxima = 0

                numero_columna = (
                    columna[0].column
                )

                letra_columna = get_column_letter(
                    numero_columna
                )

                for celda in columna:

                    if celda.value is not None:

                        longitud = len(
                            str(
                                celda.value
                            )
                        )

                        if longitud > longitud_maxima:

                            longitud_maxima = longitud

                hoja.column_dimensions[
                    letra_columna
                ].width = min(
                    longitud_maxima + 2,
                    50
                )

            # =================================================
            # GUARDAR
            # =================================================

            libro.save(
                ruta
            )

            # =================================================
            # MENSAJE
            # =================================================

            messagebox.showinfo(
                "Exportación exitosa",
                (
                    "El reporte de movimientos "
                    "se exportó correctamente "
                    "a Excel.\n\n"
                    f"Registros exportados: "
                    f"{len(movimientos)}"
                ),
                parent=self
            )

            print(
                "Excel de movimientos exportado:",
                ruta
            )

        except Exception as e:

            print(
                "Error al exportar movimientos:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudo exportar "
                    "el archivo:\n\n"
                    f"{e}"
                ),
                parent=self
            )


    # =====================================================
    # ABRIR NUEVO MOVIMIENTO
    # =====================================================

    def abrir_nuevo(self):

        # =================================================
        # EVITAR MÚLTIPLES FORMULARIOS
        # =================================================

        if (
            self.ventana_formulario is not None
            and self.ventana_formulario.winfo_exists()
        ):

            self.ventana_formulario.lift()

            self.ventana_formulario.focus_force()

            return

        try:

            # =================================================
            # IMPORTAR FORMULARIO
            # =================================================

            from app.views.formulario_movimiento import (
                FormularioMovimiento
            )

            # =================================================
            # CREAR FORMULARIO
            # =================================================

            self.ventana_formulario = FormularioMovimiento(
                self,
                callback=self.cargar_movimientos
            )

            # =================================================
            # CONFIGURAR
            # =================================================

            self.ventana_formulario.transient(
                self
            )

            self.ventana_formulario.lift()

            self.ventana_formulario.focus_force()

            self.ventana_formulario.grab_set()

            self.ventana_formulario.protocol(
                "WM_DELETE_WINDOW",
                self.cerrar_formulario
            )

        except Exception as e:

            self.ventana_formulario = None

            print(
                "Error al abrir formulario:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudo abrir el formulario "
                    "de movimiento:\n\n"
                    f"{e}"
                ),
                parent=self
            )


    # =====================================================
    # CERRAR FORMULARIO
    # =====================================================

    def cerrar_formulario(self):

        try:

            if (
                self.ventana_formulario is not None
                and self.ventana_formulario.winfo_exists()
            ):

                self.ventana_formulario.grab_release()

                self.ventana_formulario.destroy()

        except Exception as e:

            print(
                "Error al cerrar formulario:",
                e
            )

        finally:

            self.ventana_formulario = None

            try:

                self.lift()

                self.focus_force()

            except Exception:
                pass
