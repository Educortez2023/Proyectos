import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.models.responsable_model import ResponsableModel


class VentanaResponsables(ctk.CTkToplevel):

    def __init__(self, parent):

        super().__init__(parent)

        # =====================================================
        # CONFIGURACIÓN DE VENTANA
        # =====================================================

        self.parent = parent

        self.title("Gestión de Responsables")

        self.geometry("1350x700")
        self.minsize(1000, 600)

        self.transient(parent)

        # =====================================================
        # MODELO
        # =====================================================

        self.model = ResponsableModel()

        # =====================================================
        # REFERENCIA AL FORMULARIO
        # =====================================================

        self.ventana_formulario = None

        # =====================================================
        # CREAR INTERFAZ
        # =====================================================

        self.crear_interfaz()

        # =====================================================
        # AJUSTAR VENTANA
        # =====================================================

        self.after(100, self.ajustar_ventana)

        # =====================================================
        # CARGAR DATOS
        # =====================================================

        self.after(150, self.cargar_responsables)

        # =====================================================
        # MOSTRAR DELANTE
        # =====================================================

        self.after(200, self.mostrar_delante)

    # =========================================================
    # AJUSTAR VENTANA AL TAMAÑO DE LA PANTALLA
    # =========================================================

    def ajustar_ventana(self):

        try:

            ancho_pantalla = self.winfo_screenwidth()
            alto_pantalla = self.winfo_screenheight()

            margen_ancho = 80
            margen_alto = 100

            ancho = min(
                1350,
                ancho_pantalla - margen_ancho
            )

            alto = min(
                700,
                alto_pantalla - margen_alto
            )

            ancho = max(ancho, 1000)
            alto = max(alto, 600)

            x = (
                ancho_pantalla - ancho
            ) // 2

            y = (
                alto_pantalla - alto
            ) // 2

            self.geometry(
                f"{ancho}x{alto}+{x}+{y}"
            )

        except Exception as e:

            print(
                "Error al ajustar ventana:",
                e
            )

    # =========================================================
    # MOSTRAR VENTANA DELANTE
    # =========================================================

    def mostrar_delante(self):

        try:

            self.lift()
            self.focus_force()

        except Exception:
            pass

    # =========================================================
    # CREAR INTERFAZ
    # =========================================================

    def crear_interfaz(self):

        # =====================================================
        # CONFIGURACIÓN PRINCIPAL
        # =====================================================

        self.grid_rowconfigure(
            2,
            weight=1
        )

        self.grid_columnconfigure(
            0,
            weight=1
        )

        # =====================================================
        # TÍTULO
        # =====================================================

        titulo = ctk.CTkLabel(
            self,
            text="Gestión de Responsables",
            font=("Arial", 24, "bold")
        )

        titulo.grid(
            row=0,
            column=0,
            padx=20,
            pady=(20, 10),
            sticky="w"
        )

        # =====================================================
        # FRAME DE CONTROLES
        # =====================================================

        frame_controles = ctk.CTkFrame(
            self
        )

        frame_controles.grid(
            row=1,
            column=0,
            padx=20,
            pady=(5, 10),
            sticky="ew"
        )

        # =====================================================
        # CONFIGURACIÓN GRID CONTROLES
        # =====================================================

        frame_controles.grid_columnconfigure(
            6,
            weight=1
        )

        # =====================================================
        # BOTÓN NUEVO
        # =====================================================

        self.btn_nuevo = ctk.CTkButton(
            frame_controles,
            text="➕ Nuevo Responsable",
            width=170,
            height=36,
            command=self.abrir_nuevo
        )

        self.btn_nuevo.grid(
            row=0,
            column=0,
            padx=5,
            pady=10
        )

        # =====================================================
        # BOTÓN EDITAR
        # =====================================================

        self.btn_editar = ctk.CTkButton(
            frame_controles,
            text="✏ Editar",
            width=110,
            height=36,
            command=self.editar_responsable
        )

        self.btn_editar.grid(
            row=0,
            column=1,
            padx=5,
            pady=10
        )

        # =====================================================
        # BOTÓN ELIMINAR
        # =====================================================

        self.btn_eliminar = ctk.CTkButton(
            frame_controles,
            text="🗑 Eliminar",
            width=120,
            height=36,
            command=self.eliminar_responsable
        )

        self.btn_eliminar.grid(
            row=0,
            column=2,
            padx=5,
            pady=10
        )

        # =====================================================
        # BOTÓN ACTUALIZAR
        # =====================================================

        self.btn_actualizar = ctk.CTkButton(
            frame_controles,
            text="🔄 Actualizar",
            width=120,
            height=36,
            command=self.cargar_responsables
        )

        self.btn_actualizar.grid(
            row=0,
            column=3,
            padx=5,
            pady=10
        )

        # =====================================================
        # BOTÓN EXPORTAR
        # =====================================================

        self.btn_exportar = ctk.CTkButton(
            frame_controles,
            text="📊 Exportar a Excel",
            width=150,
            height=36,
            command=self.exportar_excel
        )

        self.btn_exportar.grid(
            row=0,
            column=4,
            padx=5,
            pady=10
        )

        # =====================================================
        # BOTÓN BUSCAR
        # =====================================================

        self.btn_buscar = ctk.CTkButton(
            frame_controles,
            text="🔍 Buscar",
            width=105,
            height=36,
            command=self.buscar_responsables
        )

        self.btn_buscar.grid(
            row=0,
            column=5,
            padx=5,
            pady=10
        )

        # =====================================================
        # CAMPO DE BÚSQUEDA
        # =====================================================

        self.entry_busqueda = ctk.CTkEntry(
            frame_controles,
            height=36,
            placeholder_text=(
                "Buscar por nombre, cédula, "
                "correo, teléfono..."
            )
        )

        self.entry_busqueda.grid(
            row=0,
            column=6,
            padx=(5, 10),
            pady=10,
            sticky="ew"
        )

        # =====================================================
        # ENTER PARA BUSCAR
        # =====================================================

        self.entry_busqueda.bind(
            "<Return>",
            lambda event: self.buscar_responsables()
        )

        # =====================================================
        # FRAME DE TABLA
        # =====================================================

        self.frame_tabla = ctk.CTkFrame(
            self
        )

        self.frame_tabla.grid(
            row=2,
            column=0,
            padx=20,
            pady=(0, 20),
            sticky="nsew"
        )

        # =====================================================
        # CONFIGURACIÓN GRID TABLA
        # =====================================================

        self.frame_tabla.grid_rowconfigure(
            0,
            weight=1
        )

        self.frame_tabla.grid_columnconfigure(
            0,
            weight=1
        )

        # =====================================================
        # SCROLLBAR VERTICAL
        # =====================================================

        scrollbar_vertical = ttk.Scrollbar(
            self.frame_tabla,
            orient="vertical"
        )

        scrollbar_vertical.grid(
            row=0,
            column=1,
            sticky="ns"
        )

        # =====================================================
        # SCROLLBAR HORIZONTAL
        # =====================================================

        scrollbar_horizontal = ttk.Scrollbar(
            self.frame_tabla,
            orient="horizontal"
        )

        scrollbar_horizontal.grid(
            row=1,
            column=0,
            sticky="ew"
        )

        # =====================================================
        # COLUMNAS
        # =====================================================

        columnas = (
            "id",
            "nombres",
            "apellidos",
            "cedula",
            "correo",
            "telefono",
            "departamento",
            "cargo",
            "estado"
        )

        self.columnas = columnas

        # =====================================================
        # CREAR TREEVIEW
        # =====================================================

        self.tabla = ttk.Treeview(
            self.frame_tabla,
            columns=columnas,
            show="headings",
            yscrollcommand=scrollbar_vertical.set,
            xscrollcommand=scrollbar_horizontal.set
        )

        self.tabla.grid(
            row=0,
            column=0,
            sticky="nsew"
        )

        # =====================================================
        # CONECTAR SCROLLBAR VERTICAL
        # =====================================================

        scrollbar_vertical.config(
            command=self.tabla.yview
        )

        # =====================================================
        # CONECTAR SCROLLBAR HORIZONTAL
        # =====================================================

        scrollbar_horizontal.config(
            command=self.tabla.xview
        )

        # =====================================================
        # ENCABEZADOS
        # =====================================================

        self.tabla.heading(
            "id",
            text="ID"
        )

        self.tabla.heading(
            "nombres",
            text="Nombres"
        )

        self.tabla.heading(
            "apellidos",
            text="Apellidos"
        )

        self.tabla.heading(
            "cedula",
            text="Documento / Cédula"
        )

        self.tabla.heading(
            "correo",
            text="Correo"
        )

        self.tabla.heading(
            "telefono",
            text="Teléfono"
        )

        self.tabla.heading(
            "departamento",
            text="Departamento"
        )

        self.tabla.heading(
            "cargo",
            text="Cargo"
        )

        self.tabla.heading(
            "estado",
            text="Estado"
        )

        # =====================================================
        # CONFIGURACIÓN DE COLUMNAS
        #
        # El ancho total es deliberadamente mayor que el
        # mínimo de la ventana para que el scroll horizontal
        # sea funcional cuando sea necesario.
        # =====================================================

        self.tabla.column(
            "id",
            width=60,
            minwidth=50,
            stretch=False,
            anchor="center"
        )

        self.tabla.column(
            "nombres",
            width=170,
            minwidth=120,
            stretch=True
        )

        self.tabla.column(
            "apellidos",
            width=170,
            minwidth=120,
            stretch=True
        )

        self.tabla.column(
            "cedula",
            width=150,
            minwidth=120,
            stretch=True,
            anchor="center"
        )

        self.tabla.column(
            "correo",
            width=240,
            minwidth=170,
            stretch=True
        )

        self.tabla.column(
            "telefono",
            width=130,
            minwidth=100,
            stretch=True,
            anchor="center"
        )

        self.tabla.column(
            "departamento",
            width=200,
            minwidth=140,
            stretch=True
        )

        self.tabla.column(
            "cargo",
            width=200,
            minwidth=140,
            stretch=True
        )

        self.tabla.column(
            "estado",
            width=100,
            minwidth=90,
            stretch=False,
            anchor="center"
        )

        # =====================================================
        # DOBLE CLIC PARA EDITAR
        # =====================================================

        self.tabla.bind(
            "<Double-1>",
            lambda event: self.editar_responsable()
        )

    # =========================================================
    # CARGAR RESPONSABLES
    # =========================================================

    def cargar_responsables(self):

        try:

            # =================================================
            # LIMPIAR TABLA
            # =================================================

            for item in self.tabla.get_children():

                self.tabla.delete(
                    item
                )

            # =================================================
            # OBTENER DATOS DESDE LA BD
            # =================================================

            responsables = self.model.listar()

            # =================================================
            # INSERTAR DATOS
            # =================================================

            for responsable in responsables:

                estado = (
                    "Activo"
                    if responsable[10]
                    else "Inactivo"
                )

                self.tabla.insert(
                    "",
                    "end",
                    values=(
                        responsable[0],
                        responsable[1],
                        responsable[2],
                        responsable[3],
                        responsable[4],
                        responsable[5],
                        responsable[6],
                        responsable[7],
                        estado
                    )
                )

            print(
                "Responsables cargados:",
                len(responsables)
            )

        except Exception as e:

            print(
                "Error al cargar responsables:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudieron cargar "
                    "los responsables:\n\n"
                    f"{e}"
                ),
                parent=self
            )

    # =========================================================
    # BUSCAR RESPONSABLES
    # =========================================================

    def buscar_responsables(self):

        try:

            texto = (
                self.entry_busqueda
                .get()
                .strip()
            )

            # =================================================
            # SI ESTÁ VACÍO
            # =================================================

            if not texto:

                self.cargar_responsables()

                return

            # =================================================
            # LIMPIAR TABLA
            # =================================================

            for item in self.tabla.get_children():

                self.tabla.delete(
                    item
                )

            # =================================================
            # BUSCAR EN BD
            # =================================================

            responsables = self.model.buscar(
                texto
            )

            # =================================================
            # INSERTAR RESULTADOS
            # =================================================

            for responsable in responsables:

                estado = (
                    "Activo"
                    if responsable[10]
                    else "Inactivo"
                )

                self.tabla.insert(
                    "",
                    "end",
                    values=(
                        responsable[0],
                        responsable[1],
                        responsable[2],
                        responsable[3],
                        responsable[4],
                        responsable[5],
                        responsable[6],
                        responsable[7],
                        estado
                    )
                )

            print(
                "Responsables encontrados:",
                len(responsables)
            )

        except Exception as e:

            print(
                "Error al buscar responsables:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudo realizar "
                    "la búsqueda:\n\n"
                    f"{e}"
                ),
                parent=self
            )

    # =========================================================
    # ABRIR NUEVO RESPONSABLE
    # =========================================================

    def abrir_nuevo(self):

        # =====================================================
        # EVITAR MÚLTIPLES FORMULARIOS
        # =====================================================

        if (
            self.ventana_formulario is not None
            and self.ventana_formulario.winfo_exists()
        ):

            self.ventana_formulario.lift()
            self.ventana_formulario.focus_force()

            return

        try:

            from app.views.formulario_responsable import (
                FormularioResponsable
            )

            self.ventana_formulario = (
                FormularioResponsable(
                    self,
                    callback=self.cargar_responsables
                )
            )

            self.ventana_formulario.transient(
                self
            )

            self.ventana_formulario.lift()
            self.ventana_formulario.focus_force()

            self.ventana_formulario.grab_set()

            self.ventana_formulario.protocol(
                "WM_DELETE_WINDOW",
                self.cerrar_formulario
            )

        except Exception as e:

            self.ventana_formulario = None

            print(
                "Error al abrir formulario:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudo abrir el formulario "
                    "de responsable:\n\n"
                    f"{e}"
                ),
                parent=self
            )

    # =========================================================
    # CERRAR FORMULARIO
    # =========================================================

    def cerrar_formulario(self):

        try:

            if (
                self.ventana_formulario is not None
                and self.ventana_formulario.winfo_exists()
            ):

                try:

                    self.ventana_formulario.grab_release()

                except Exception:
                    pass

                self.ventana_formulario.destroy()

        except Exception as e:

            print(
                "Error al cerrar formulario:",
                e
            )

        finally:

            self.ventana_formulario = None

            try:

                self.lift()
                self.focus_force()

            except Exception:
                pass

    # =========================================================
    # EDITAR RESPONSABLE
    # =========================================================

    def editar_responsable(self):

        seleccionado = self.tabla.selection()

        # =====================================================
        # VALIDAR SELECCIÓN
        # =====================================================

        if not seleccionado:

            messagebox.showwarning(
                "Advertencia",
                (
                    "Seleccione un responsable "
                    "para editar."
                ),
                parent=self
            )

            return

        # =====================================================
        # VERIFICAR FORMULARIO
        # =====================================================

        if (
            self.ventana_formulario is not None
            and self.ventana_formulario.winfo_exists()
        ):

            self.ventana_formulario.lift()
            self.ventana_formulario.focus_force()

            return

        # =====================================================
        # OBTENER ID
        # =====================================================

        item = seleccionado[0]

        valores = self.tabla.item(
            item,
            "values"
        )

        if not valores:

            messagebox.showerror(
                "Error",
                (
                    "No se pudieron obtener "
                    "los datos del responsable."
                ),
                parent=self
            )

            return

        id_responsable = valores[0]

        try:

            # =================================================
            # OBTENER DATOS COMPLETOS
            # =================================================

            responsables = self.model.listar()

            responsable_completo = None

            for responsable in responsables:

                if str(
                    responsable[0]
                ) == str(
                    id_responsable
                ):

                    responsable_completo = responsable

                    break

            # =================================================
            # VALIDAR
            # =================================================

            if responsable_completo is None:

                messagebox.showerror(
                    "Error",
                    (
                        "No se encontró la información "
                        "completa del responsable."
                    ),
                    parent=self
                )

                return

            # =================================================
            # IMPORTAR FORMULARIO
            # =================================================

            from app.views.formulario_responsable import (
                FormularioResponsable
            )

            # =================================================
            # ABRIR FORMULARIO
            # =================================================

            self.ventana_formulario = (
                FormularioResponsable(
                    self,
                    responsable=responsable_completo,
                    callback=self.cargar_responsables
                )
            )

            self.ventana_formulario.transient(
                self
            )

            self.ventana_formulario.lift()
            self.ventana_formulario.focus_force()

            self.ventana_formulario.grab_set()

            self.ventana_formulario.protocol(
                "WM_DELETE_WINDOW",
                self.cerrar_formulario
            )

        except Exception as e:

            self.ventana_formulario = None

            print(
                "Error al editar responsable:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudo abrir el formulario "
                    "de edición:\n\n"
                    f"{e}"
                ),
                parent=self
            )

    # =========================================================
    # ELIMINAR RESPONSABLE
    # =========================================================

    def eliminar_responsable(self):

        seleccionado = self.tabla.selection()

        if not seleccionado:

            messagebox.showwarning(
                "Advertencia",
                (
                    "Seleccione un responsable "
                    "para eliminar."
                ),
                parent=self
            )

            return

        valores = self.tabla.item(
            seleccionado[0],
            "values"
        )

        if not valores:

            messagebox.showerror(
                "Error",
                (
                    "No se pudieron obtener "
                    "los datos del responsable."
                ),
                parent=self
            )

            return

        id_responsable = valores[0]

        nombre_completo = (
            f"{valores[1]} {valores[2]}"
        )

        confirmar = messagebox.askyesno(
            "Confirmar eliminación",
            (
                "¿Está seguro de eliminar "
                f"al responsable:\n\n"
                f"{nombre_completo}?"
            ),
            parent=self
        )

        if not confirmar:

            return

        try:

            self.model.eliminar(
                id_responsable
            )

            messagebox.showinfo(
                "Éxito",
                (
                    "Responsable eliminado "
                    "correctamente."
                ),
                parent=self
            )

            self.cargar_responsables()

        except Exception as e:

            print(
                "Error al eliminar responsable:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudo eliminar "
                    "el responsable:\n\n"
                    f"{e}"
                ),
                parent=self
            )

    # =========================================================
    # EXPORTAR A EXCEL
    # =========================================================

    def exportar_excel(self):

        try:

            # =================================================
            # OBTENER DIRECTAMENTE DE LA BD
            # =================================================

            responsables = self.model.listar()

            if not responsables:

                messagebox.showwarning(
                    "Sin datos",
                    (
                        "No existen responsables "
                        "registrados para exportar."
                    ),
                    parent=self
                )

                return

            # =================================================
            # SELECCIONAR UBICACIÓN
            # =================================================

            ruta = filedialog.asksaveasfilename(
                parent=self,
                title="Guardar reporte de responsables",
                defaultextension=".xlsx",
                filetypes=[
                    (
                        "Archivo Excel",
                        "*.xlsx"
                    ),
                    (
                        "Todos los archivos",
                        "*.*"
                    )
                ],
                initialfile="Reporte_Responsables.xlsx"
            )

            if not ruta:

                return

            # =================================================
            # CREAR LIBRO
            # =================================================

            libro = Workbook()

            hoja = libro.active

            hoja.title = "Responsables"

            # =================================================
            # ENCABEZADOS
            # =================================================

            encabezados = [
                "ID",
                "Nombres",
                "Apellidos",
                "Documento / Cédula",
                "Correo",
                "Teléfono",
                "Departamento",
                "Cargo",
                "Estado"
            ]

            hoja.append(
                encabezados
            )

            # =================================================
            # FORMATO
            # =================================================

            for celda in hoja[1]:

                celda.font = Font(
                    bold=True
                )

            # =================================================
            # DATOS
            # =================================================

            for responsable in responsables:

                estado = (
                    "Activo"
                    if responsable[10]
                    else "Inactivo"
                )

                hoja.append(
                    [
                        responsable[0],
                        responsable[1],
                        responsable[2],
                        responsable[3],
                        responsable[4],
                        responsable[5],
                        responsable[6],
                        responsable[7],
                        estado
                    ]
                )

            # =================================================
            # FILTRO
            # =================================================

            ultima_fila = hoja.max_row

            ultima_columna = hoja.max_column

            ultima_columna_letra = get_column_letter(
                ultima_columna
            )

            hoja.auto_filter.ref = (
                f"A1:{ultima_columna_letra}{ultima_fila}"
            )

            # =================================================
            # CONGELAR ENCABEZADOS
            # =================================================

            hoja.freeze_panes = "A2"

            # =================================================
            # AJUSTAR COLUMNAS
            # =================================================

            for columna in hoja.columns:

                longitud_maxima = 0

                numero_columna = (
                    columna[0].column
                )

                letra_columna = get_column_letter(
                    numero_columna
                )

                for celda in columna:

                    if celda.value is not None:

                        longitud = len(
                            str(
                                celda.value
                            )
                        )

                        if longitud > longitud_maxima:

                            longitud_maxima = longitud

                hoja.column_dimensions[
                    letra_columna
                ].width = min(
                    longitud_maxima + 2,
                    50
                )

            # =================================================
            # GUARDAR
            # =================================================

            libro.save(
                ruta
            )

            # =================================================
            # MENSAJE
            # =================================================

            messagebox.showinfo(
                "Exportación exitosa",
                (
                    "El reporte de responsables "
                    "se exportó correctamente "
                    "a Excel.\n\n"
                    f"Registros exportados: "
                    f"{len(responsables)}"
                ),
                parent=self
            )

            print(
                "Excel de responsables exportado:",
                ruta
            )

        except Exception as e:

            print(
                "Error al exportar responsables:",
                e
            )

            messagebox.showerror(
                "Error",
                (
                    "No se pudo exportar "
                    "el archivo:\n\n"
                    f"{e}"
                ),
                parent=self
            )